import io
import os
import sqlite3
import tempfile
import unittest
from contextlib import closing

import api
import core
from openpyxl import load_workbook
from services.request_context import set_request_identity


class HistorySyncTest(unittest.TestCase):
    def setUp(self):
        set_request_identity("local", guest=False)
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_file = os.path.join(self.temp_dir.name, "test.db")
        os.environ["DB_PATH"] = self.db_file
        core.init_db()
        api.STATE.update({
            "records": [],
            "source_columns": [],
            "id_column": None,
            "display_columns": [],
            "special_rules": [],
            "excluded_indices": set(),
            "last_winner_indices": set(),
            "latest_session_id": None,
            "csv_file": "",
            "event_id": None,
            "user_event_id": "__default__",
            "history_import": None,
        })
        self.a_id = self._seed("A", 2, 1)
        self.c_id = self._seed("C", 4, 0)

    def tearDown(self):
        set_request_identity("local", guest=False)
        os.environ.pop("DB_PATH", None)
        self.temp_dir.cleanup()

    def _seed(self, draw_id, join_count, win_count):
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            participant_id = conn.execute(
                "INSERT INTO participants (display_name) VALUES (?)", (draw_id,)
            ).lastrowid
            conn.execute("""
                INSERT INTO event_participant_history
                (event_id, participant_id, join_count, win_count, streak_count)
                VALUES (0, ?, ?, ?, ?)
            """, (participant_id, join_count, win_count, join_count - win_count))
        return participant_id

    def _rows(self):
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            return {
                row[0]: (row[1], row[2])
                for row in conn.execute("""
                    SELECT p.display_name, h.join_count, h.win_count
                    FROM event_participant_history h
                    JOIN participants p ON p.id=h.participant_id
                    WHERE h.event_id=0
                """)
            }

    def _add_submission(self, participant_id):
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            session_id = conn.execute("""
                INSERT INTO raffle_sessions (event_id, created_at)
                VALUES (NULL, '2026-07-13T00:00:00')
            """).lastrowid
            conn.execute("""
                INSERT INTO submission_records
                (session_id, matched_participant_id, created_at)
                VALUES (?, ?, '2026-07-13T00:00:00')
            """, (session_id, participant_id))

    def _import(self, rows, mode=None):
        api.STATE["history_import"] = {
            "filename": "history.csv",
            "columns": ["ID", "参加", "当選"],
            "rows": rows,
        }
        payload = {
            "eventId": "__default__",
            "idColumn": "ID",
            "joinColumn": "参加",
            "winColumn": "当選",
        }
        if mode:
            payload["syncMode"] = mode
        api.handle_history_apply(payload)
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            return conn.execute(
                "SELECT max(id) FROM history_sync_batches"
            ).fetchone()[0]

    def test_add_is_default_and_can_be_rolled_back(self):
        batch_id = self._import([
            {"ID": "A", "参加": 3, "当選": 1},
            {"ID": "B", "参加": 1, "当選": 0},
        ])
        self.assertEqual(self._rows(), {"A": (5, 2), "B": (1, 0), "C": (4, 0)})

        api.handle_history_rollback({"batchId": batch_id})
        self.assertEqual(self._rows(), {"A": (2, 1), "C": (4, 0)})

    def test_overwrite_replaces_entire_event_and_rollback_restores_it(self):
        self._add_submission(self.c_id)
        batch_id = self._import([
            {"ID": "B", "参加": 7, "当選": 2},
        ], "overwrite")
        self.assertEqual(self._rows(), {"B": (7, 2)})
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            snapshot_complete = conn.execute(
                "SELECT snapshot_complete FROM history_sync_batches WHERE id=?",
                (batch_id,),
            ).fetchone()[0]
            snapshot_count = conn.execute(
                "SELECT count(*) FROM history_sync_snapshots WHERE batch_id=?",
                (batch_id,),
            ).fetchone()[0]
        self.assertEqual(snapshot_complete, 1)
        self.assertEqual(snapshot_count, 2)
        self.assertEqual(
            [row["drawId"] for row in api.public_state()["savedUsers"]],
            ["B"],
        )

        api.handle_history_rollback({"batchId": batch_id})
        self.assertEqual(self._rows(), {"A": (2, 1), "C": (4, 0)})

    def test_selecting_existing_id_column_works_with_default_sqlite_rows(self):
        self._add_submission(self.a_id)

        api.STATE["source_columns"] = ["ID", "名前"]
        api.STATE["records"] = [{
            "raw_values": {"ID": "A", "名前": "Alice"},
            "base_display_name": "Alice",
        }]
        result = api.handle_roles({"idColumn": "ID", "displayColumns": []})

        self.assertTrue(result["ok"])
        self.assertEqual(result["idColumn"], "ID")
        self.assertEqual(api.STATE["records"][0]["join_count"], 2)

    def test_raffle_updates_authoritative_event_user_table(self):
        api.STATE.update({
            "source_columns": ["ID"],
            "id_column": "ID",
            "records": [{
                "raw_values": {"ID": "A"},
                "draw_id": "A",
                "display_name": "A",
                "base_display_name": "A",
                "display_fields": {},
                "participant_id": self.a_id,
                "matched": True,
                "join_count": 2,
                "win_count": 1,
                "streak_count": 1,
            }],
            "csv_file": "raffle.csv",
            "mode": "equal",
        })

        api.run_raffle({"drawCount": 1, "mode": "equal", "eventId": ""})

        self.assertEqual(self._rows()["A"], (3, 2))

    def test_deleting_session_rolls_back_user_statistics(self):
        api.STATE.update({
            "source_columns": ["ID"],
            "id_column": "ID",
            "records": [{
                "raw_values": {"ID": "A"},
                "draw_id": "A",
                "display_name": "A",
                "base_display_name": "A",
                "display_fields": {},
                "participant_id": self.a_id,
                "matched": True,
                "join_count": 2,
                "win_count": 1,
                "streak_count": 1,
            }],
            "csv_file": "raffle.csv",
            "mode": "equal",
            "event_id": None,
            "user_event_id": "__default__",
        })
        result = api.run_raffle({"drawCount": 1, "mode": "equal", "eventId": ""})
        session_id = result["latestSessionId"]

        deleted = api.handle_action("/api/session/delete", {"sessionId": session_id})

        self.assertEqual(self._rows()["A"], (2, 1))
        saved = next(row for row in deleted["savedUsers"] if row["drawId"] == "A")
        self.assertEqual((saved["join_count"], saved["win_count"]), (2, 1))

    def test_default_event_can_be_renamed_and_requires_another_event_to_delete(self):
        renamed = api.handle_event_save({
            "eventId": "__default__",
            "name": "メインイベント",
            "description": "",
        })
        self.assertEqual(renamed["defaultEventName"], "メインイベント")

        with self.assertRaisesRegex(ValueError, "他のEvent"):
            api.handle_event_delete({"eventId": "__default__"})

        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            target_id = conn.execute("""
                INSERT INTO events (name, created_at) VALUES ('Event 2', '2026-07-13')
            """).lastrowid
        self._add_submission(self.a_id)
        deleted = api.handle_event_delete({
            "eventId": "__default__",
            "targetEventId": target_id,
        })

        self.assertFalse(deleted["defaultEventEnabled"])
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            self.assertEqual(conn.execute(
                "SELECT count(*) FROM event_participant_history WHERE event_id=0"
            ).fetchone()[0], 0)
            self.assertEqual(conn.execute(
                "SELECT count(*) FROM event_participant_history WHERE event_id=?",
                (target_id,),
            ).fetchone()[0], 2)
            self.assertEqual(conn.execute(
                "SELECT count(*) FROM raffle_sessions WHERE event_id=?",
                (target_id,),
            ).fetchone()[0], 1)

    def test_excel_export_contains_user_probability_and_raffle_sheets(self):
        with closing(sqlite3.connect(self.db_file)) as conn, conn:
            session_id = conn.execute("""
                INSERT INTO raffle_sessions
                (event_id, csv_file, mode, draw_count, created_at)
                VALUES (NULL, 'sample.csv', 'linear', 1, '2026-07-13T10:30:00')
            """).lastrowid
            conn.execute("""
                INSERT INTO raffle_results
                (session_id, participant_id, display_name, is_winner)
                VALUES (?, ?, 'A', 1)
            """, (session_id, self.a_id))

        content, filename = api.build_event_export("__default__")
        workbook = load_workbook(io.BytesIO(content), data_only=True)

        self.assertEqual(workbook.sheetnames, ["Sheet1", "Sheet2"])
        self.assertTrue(filename.endswith(".xlsx"))
        self.assertEqual(workbook["Sheet1"]["A2"].value, "A")
        self.assertIsInstance(workbook["Sheet1"]["E2"].value, float)
        self.assertEqual(workbook["Sheet2"]["D2"].value, "A")

    def test_saved_events_are_isolated_by_x_account(self):
        set_request_identity("x:account-a", guest=False)
        api.handle_event_save({"name": "A Event", "description": ""})
        self.assertEqual([event["name"] for event in api.public_state()["events"]], ["A Event"])

        set_request_identity("x:account-b", guest=False)
        self.assertEqual(api.public_state()["events"], [])

    def test_guest_raffle_does_not_create_a_session(self):
        set_request_identity("guest:test", guest=True)
        api.STATE.update({
            "source_columns": ["ID"],
            "id_column": "ID",
            "records": [{
                "raw_values": {"ID": "Guest"}, "draw_id": "Guest",
                "display_name": "Guest", "base_display_name": "Guest",
                "display_fields": {}, "participant_id": None, "matched": False,
                "join_count": 0, "win_count": 0, "streak_count": 0,
            }],
            "csv_file": "guest.csv",
            "mode": "equal",
        })

        result = api.run_raffle({"drawCount": 1, "mode": "double", "eventId": ""})

        self.assertTrue(result["guestMode"])
        self.assertEqual(result["mode"], "equal")
        with closing(sqlite3.connect(self.db_file)) as conn:
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM raffle_sessions").fetchone()[0], 0)


if __name__ == "__main__":
    unittest.main()
